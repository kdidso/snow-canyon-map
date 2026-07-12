from __future__ import annotations

import os
import re
import sys
import time
from datetime import date, datetime
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException, StaleElementReferenceException, TimeoutException
from selenium.webdriver import ChromeOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select, WebDriverWait


# ============================================================
# CONFIG
# ============================================================

LCR_BASE = "https://lcr.churchofjesuschrist.org"
ATTENDANCE_PAGE_URL = (
    f"{LCR_BASE}/mlt/report/class-and-quorum-attendance?lang=eng"
)

USERNAME = os.getenv("LCR_USERNAME", "").strip()
PASSWORD = os.getenv("LCR_PASSWORD", "").strip()

START_DATE = os.getenv("START_DATE", "2025-12-28").strip()
END_DATE = os.getenv("END_DATE", "2026-03-08").strip()
OUTPUT_DIR = Path("data")

HEADLESS = True
LONG_WAIT = 60
ATTENDANCE_WAIT = 180
MONTH_CHANGE_WAIT = 90

DEBUG_ATTENDANCE_HTML = OUTPUT_DIR / "debug_attendance_page.html"
DEBUG_ATTENDANCE_TEXT = OUTPUT_DIR / "debug_attendance_page.txt"
DEBUG_ATTENDANCE_ROWS = OUTPUT_DIR / "debug_attendance_rows.txt"
DEBUG_ATTENDANCE_CONTROLS = OUTPUT_DIR / "debug_attendance_controls.txt"
DEBUG_ATTENDANCE_SCREENSHOT = OUTPUT_DIR / "debug_attendance_page.png"


# ============================================================
# HELPERS
# ============================================================

def log(msg: str) -> None:
    print(f"[INFO] {msg}")


def err(msg: str) -> None:
    print(f"[ERROR] {msg}", file=sys.stderr)


def parse_iso_date(value: str) -> date:
    return datetime.strptime(value, "%Y-%m-%d").date()


def format_excel_header(dt: date) -> str:
    return f"{dt.strftime('%b')} {dt.day} {dt.year}"


def get_body_text(driver: webdriver.Chrome) -> str:
    return driver.find_element(By.TAG_NAME, "body").text


def clean_name(name: str) -> str:
    name = re.sub(r"\s*(Out-of-Unit|Not Baptized)\s*", " ", name)
    name = re.sub(r"\s+", " ", name)
    return name.strip()


def looks_like_name(name: str) -> bool:
    if not name or name == "Come, Follow Me" or len(name) > 100:
        return False
    return bool(
        re.match(
            r"^[A-Za-zÀ-ÿ'’\.\- ]+,\s+[A-Za-zÀ-ÿ'’\.\- ]+$",
            name,
        )
    )


def month_key(dt: date) -> str:
    return dt.strftime("%Y-%m")


def month_keys_between(start_dt: date, end_dt: date) -> List[str]:
    keys: List[str] = []
    year, month = start_dt.year, start_dt.month
    while (year, month) <= (end_dt.year, end_dt.month):
        keys.append(f"{year:04d}-{month:02d}")
        month += 1
        if month == 13:
            month = 1
            year += 1
    return keys


def parse_col_date(col_id: str) -> Optional[date]:
    match = re.match(r"week_(\d{4}-\d{2}-\d{2})_SORT$", col_id or "")
    if not match:
        return None
    try:
        return datetime.strptime(match.group(1), "%Y-%m-%d").date()
    except ValueError:
        return None


# ============================================================
# SELENIUM LOGIN
# ============================================================

def make_driver() -> webdriver.Chrome:
    opts = ChromeOptions()
    if HEADLESS or os.getenv("CI", "").lower() == "true":
        opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--window-size=1920,3000")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_argument("--lang=en-US")
    return webdriver.Chrome(options=opts)


def login(driver: webdriver.Chrome) -> None:
    if not USERNAME or not PASSWORD:
        raise RuntimeError("Missing LCR_USERNAME and/or LCR_PASSWORD.")

    log("Opening LCR login page")
    driver.get(LCR_BASE)

    user_input = WebDriverWait(driver, LONG_WAIT).until(
        EC.presence_of_element_located((By.ID, "username-input"))
    )
    user_input.clear()
    user_input.send_keys(USERNAME)
    user_input.send_keys(Keys.ENTER)

    pwd_input = WebDriverWait(driver, LONG_WAIT).until(
        EC.presence_of_element_located((By.ID, "password-input"))
    )
    pwd_input.clear()
    pwd_input.send_keys(PASSWORD)
    pwd_input.send_keys(Keys.ENTER)

    WebDriverWait(driver, LONG_WAIT).until(EC.url_contains(LCR_BASE))
    log("Login submitted successfully")


# ============================================================
# DEBUG OUTPUT
# ============================================================

def save_debug(driver: webdriver.Chrome, row_lines: Optional[List[str]] = None) -> None:
    try:
        DEBUG_ATTENDANCE_HTML.write_text(driver.page_source, encoding="utf-8")
        DEBUG_ATTENDANCE_TEXT.write_text(get_body_text(driver), encoding="utf-8")
    except Exception:
        pass

    try:
        driver.save_screenshot(str(DEBUG_ATTENDANCE_SCREENSHOT))
    except Exception as ex:
        log(f"Could not save attendance screenshot: {ex}")

    controls: List[str] = []
    try:
        for index, el in enumerate(
            driver.find_elements(By.CSS_SELECTOR, "input, button, select"), start=1
        ):
            try:
                controls.append(
                    " | ".join(
                        [
                            str(index),
                            f"tag={el.tag_name}",
                            f"type={el.get_attribute('type') or ''}",
                            f"id={el.get_attribute('id') or ''}",
                            f"title={el.get_attribute('title') or ''}",
                            f"value={el.get_attribute('value') or ''}",
                            f"text={re.sub(r'\\s+', ' ', el.text or '').strip()}",
                        ]
                    )
                )
            except StaleElementReferenceException:
                continue
    except Exception:
        pass

    DEBUG_ATTENDANCE_CONTROLS.write_text("\n".join(controls), encoding="utf-8")
    if row_lines is not None:
        DEBUG_ATTENDANCE_ROWS.write_text("\n".join(row_lines), encoding="utf-8")


# ============================================================
# PAGE AND MONTH SELECTION
# ============================================================

def wait_for_attendance_table(driver: webdriver.Chrome) -> None:
    WebDriverWait(driver, ATTENDANCE_WAIT).until(
        lambda d: len(d.find_elements(By.CSS_SELECTOR, "table[role='grid'] tbody tr")) > 5
        and len(d.find_elements(By.CSS_SELECTOR, "colgroup col[id^='week_']")) > 0
    )


def find_month_select(driver: webdriver.Chrome):
    """Find the select whose option values use YYYY-MM."""
    for select_el in driver.find_elements(By.CSS_SELECTOR, "select"):
        try:
            values = [
                option.get_attribute("value") or ""
                for option in select_el.find_elements(By.TAG_NAME, "option")
            ]
            if any(re.fullmatch(r"\d{4}-\d{2}", value) for value in values):
                return select_el
        except StaleElementReferenceException:
            continue
    raise RuntimeError("Could not find the attendance month selector.")


def visible_month_from_columns(driver: webdriver.Chrome) -> str:
    for col in driver.find_elements(By.CSS_SELECTOR, "colgroup col[id^='week_']"):
        dt = parse_col_date(col.get_attribute("id") or "")
        if dt:
            return month_key(dt)
    return ""


def select_month(driver: webdriver.Chrome, target_month: str) -> None:
    select_el = find_month_select(driver)
    selector = Select(select_el)
    available_values = {
        option.get_attribute("value") or "" for option in selector.options
    }

    if target_month not in available_values:
        raise RuntimeError(
            f"Attendance month {target_month} is not available in the LCR month selector. "
            f"Available months include: {', '.join(sorted(v for v in available_values if re.fullmatch(r'\\d{{4}}-\\d{{2}}', v)))}"
        )

    current_value = select_el.get_attribute("value") or ""
    if current_value == target_month and visible_month_from_columns(driver) == target_month:
        return

    old_signature = tuple(
        col.get_attribute("id") or ""
        for col in driver.find_elements(By.CSS_SELECTOR, "colgroup col[id^='week_']")
    )

    log(f"Selecting attendance month: {target_month}")
    selector.select_by_value(target_month)

    def month_changed(d: webdriver.Chrome) -> bool:
        try:
            new_signature = tuple(
                col.get_attribute("id") or ""
                for col in d.find_elements(By.CSS_SELECTOR, "colgroup col[id^='week_']")
            )
            return (
                bool(new_signature)
                and new_signature != old_signature
                and visible_month_from_columns(d) == target_month
                and len(d.find_elements(By.CSS_SELECTOR, "table[role='grid'] tbody tr")) > 5
            )
        except StaleElementReferenceException:
            return False

    WebDriverWait(driver, MONTH_CHANGE_WAIT).until(month_changed)
    time.sleep(0.5)


# ============================================================
# ATTENDANCE EXTRACTION
# ============================================================

def attendance_state_from_cell(cell) -> Optional[bool]:
    """
    Return:
      True  = checked attendance icon
      False = empty-circle attendance icon
      None  = no attendance button, usually an unavailable/future date
    """
    buttons = cell.find_elements(
        By.CSS_SELECTOR,
        "button[class*='attendanceButton']",
    )
    if not buttons:
        return None

    paths = buttons[0].find_elements(By.CSS_SELECTOR, "svg path")
    if not paths:
        return None

    path = paths[0]
    fill_rule = (path.get_attribute("fill-rule") or "").strip().lower()
    clip_rule = (path.get_attribute("clip-rule") or "").strip().lower()
    path_data = re.sub(r"\s+", " ", path.get_attribute("d") or "").strip()

    # The checked-circle SVG has evenodd rules and starts with M12 22...
    if fill_rule == "evenodd" or clip_rule == "evenodd":
        return True
    if path_data.startswith("M12 22"):
        return True

    # The unmarked icon is the outlined circle and starts with M12 3.5...
    if path_data.startswith("M12 3.5"):
        return False

    raise RuntimeError(f"Unrecognized attendance SVG path: {path_data[:120]}")


def scrape_visible_month(
    driver: webdriver.Chrome,
    requested_start: date,
    requested_end: date,
) -> Tuple[Dict[str, Dict[date, Optional[bool]]], Set[date], List[str]]:
    table = WebDriverWait(driver, ATTENDANCE_WAIT).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, "table[role='grid']"))
    )

    col_dates: List[Optional[date]] = []
    for col in table.find_elements(By.CSS_SELECTOR, "colgroup col"):
        col_dates.append(parse_col_date(col.get_attribute("id") or ""))

    if len(col_dates) < 3:
        raise RuntimeError("Attendance table did not contain expected date columns.")

    data: Dict[str, Dict[date, Optional[bool]]] = {}
    available_dates: Set[date] = set()
    debug_rows: List[str] = []

    for row_index, row in enumerate(
        table.find_elements(By.CSS_SELECTOR, "tbody tr[role='row']"), start=1
    ):
        cells = row.find_elements(By.CSS_SELECTOR, "td")
        if len(cells) < 3:
            continue

        name_buttons = cells[0].find_elements(
            By.CSS_SELECTOR, "button[data-member-card-person-uuid]"
        )
        if not name_buttons:
            continue

        name = clean_name(name_buttons[0].text)
        if not looks_like_name(name):
            continue

        uuid = name_buttons[0].get_attribute("data-member-card-person-uuid") or ""
        states_for_debug: List[str] = []
        data.setdefault(name, {})

        # col_dates[0] and col_dates[1] are Name and Gender and therefore None.
        for col_index in range(2, min(len(cells), len(col_dates))):
            dt = col_dates[col_index]
            if dt is None or not (requested_start <= dt <= requested_end):
                continue

            state = attendance_state_from_cell(cells[col_index])
            data[name][dt] = state

            if state is not None:
                available_dates.add(dt)

            states_for_debug.append(
                f"{dt.isoformat()}={'present' if state is True else 'absent' if state is False else 'unavailable'}"
            )

        debug_rows.append(
            f"ROW {row_index} | uuid={uuid} | name={name} | " + " | ".join(states_for_debug)
        )

    return data, available_dates, debug_rows


def merge_attendance(
    destination: Dict[str, Dict[date, Optional[bool]]],
    incoming: Dict[str, Dict[date, Optional[bool]]],
) -> None:
    for name, values in incoming.items():
        destination.setdefault(name, {}).update(values)


def scrape_attendance(
    driver: webdriver.Chrome,
    start_dt: date,
    end_dt: date,
) -> Tuple[Dict[str, Dict[date, Optional[bool]]], List[date]]:
    log(f"Loading rendered attendance page: {ATTENDANCE_PAGE_URL}")
    driver.get(ATTENDANCE_PAGE_URL)
    wait_for_attendance_table(driver)

    combined: Dict[str, Dict[date, Optional[bool]]] = {}
    all_available_dates: Set[date] = set()
    all_debug_rows: List[str] = []

    try:
        for target_month in month_keys_between(start_dt, end_dt):
            select_month(driver, target_month)
            month_data, month_dates, month_debug = scrape_visible_month(
                driver, start_dt, end_dt
            )
            merge_attendance(combined, month_data)
            all_available_dates.update(month_dates)
            all_debug_rows.extend([f"MONTH {target_month}"] + month_debug)
            log(
                f"Collected {len(month_dates)} available attendance dates "
                f"and {sum(1 for v in month_data.values() if v)} member rows for {target_month}"
            )
    except Exception:
        save_debug(driver, all_debug_rows)
        raise

    save_debug(driver, all_debug_rows)

    final_dates = sorted(
        dt for dt in all_available_dates if start_dt <= dt <= end_dt
    )
    people_with_values = sum(
        1
        for per_date in combined.values()
        if any(value is not None for value in per_date.values())
    )

    log(f"Attendance dates collected: {len(final_dates)}")
    log(f"Members with attendance values: {people_with_values}")

    if not final_dates or people_with_values == 0:
        raise RuntimeError(
            "The attendance page loaded, but no usable attendance values were collected. "
            f"Debug files were written under {OUTPUT_DIR}."
        )

    return combined, final_dates


# ============================================================
# EXCEL OUTPUT
# ============================================================

def write_excel(
    attendance_data: Dict[str, Dict[date, Optional[bool]]],
    all_dates: List[date],
    out_path: Path,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    percent_fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
    unavailable_fill = PatternFill(fill_type="solid", fgColor="E7E6E6")
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    headers = ["Name", "% activity"] + [format_excel_header(dt) for dt in all_dates]
    ws.append(headers)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = left if col_idx == 1 else center

    names = [
        name
        for name, values in attendance_data.items()
        if any(values.get(dt) is not None for dt in all_dates)
    ]

    for row_idx, name in enumerate(sorted(names, key=str.casefold), start=2):
        per_date = attendance_data[name]
        countable_dates = [dt for dt in all_dates if per_date.get(dt) is not None]
        present_count = sum(1 for dt in countable_dates if per_date.get(dt) is True)
        pct = present_count / len(countable_dates) if countable_dates else 0.0

        ws.cell(row=row_idx, column=1, value=name)

        pct_cell = ws.cell(row=row_idx, column=2, value=pct)
        pct_cell.number_format = "0%"
        pct_cell.fill = percent_fill

        for col_idx, dt in enumerate(all_dates, start=3):
            state = per_date.get(dt)
            cell = ws.cell(row=row_idx, column=col_idx)
            if state is True:
                cell.value = "☑"
            elif state is False:
                cell.value = "☐"
            else:
                cell.value = "—"
                cell.fill = unavailable_fill

    ws.freeze_panes = "C2"
    ws.auto_filter.ref = ws.dimensions
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 12

    for col_idx in range(3, 3 + len(all_dates)):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14

    for row in ws.iter_rows(
        min_row=2,
        max_row=ws.max_row,
        min_col=1,
        max_col=ws.max_column,
    ):
        for cell in row:
            cell.alignment = left if cell.column == 1 else center

    wb.save(out_path)


# ============================================================
# MAIN
# ============================================================

def main() -> int:
    if not USERNAME or not PASSWORD:
        err("Missing LCR_USERNAME and/or LCR_PASSWORD environment variables.")
        return 1

    try:
        start_dt = parse_iso_date(START_DATE)
        end_dt = parse_iso_date(END_DATE)
    except ValueError:
        err("START_DATE and END_DATE must use YYYY-MM-DD format.")
        return 1

    if start_dt > end_dt:
        err("START_DATE must be on or before END_DATE.")
        return 1

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = OUTPUT_DIR / (
        f"attendance_{start_dt.isoformat()}_to_{end_dt.isoformat()}_{timestamp}.xlsx"
    )

    driver = make_driver()
    try:
        login(driver)
        attendance_data, all_dates = scrape_attendance(driver, start_dt, end_dt)
        write_excel(attendance_data, all_dates, out_path)
        log(f"Excel output written to {out_path}")
        return 0
    except Exception as ex:
        err(str(ex))
        try:
            save_debug(driver)
        except Exception:
            pass
        return 1
    finally:
        try:
            driver.quit()
        except Exception:
            pass


if __name__ == "__main__":
    sys.exit(main())
